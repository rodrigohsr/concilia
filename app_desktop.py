"""Concilia - conferencia de extratos bancarios em OFX.

Interface desktop (Tkinter) usada no fluxo da Esquema Assessoria Contabil.

Notas de desempenho: nada pesado e importado na inicializacao. O parser usa
apenas a biblioteca padrao, e reportlab/openpyxl/PIL sao carregados sob demanda,
na primeira exportacao. Isso mantem a abertura do programa praticamente
instantanea, inclusive no executavel gerado pelo PyInstaller.
"""

from __future__ import annotations

import json
import os
import sys
import tkinter as tk
from dataclasses import dataclass
from tkinter import filedialog, messagebox, ttk

from ofx_parser import ExtratoOFX, OFXParserBR

APP_NOME = "Concilia"
VERSAO = "1.1.0"
DESCRICAO = "Conferencia de extratos bancarios em OFX"
EMPRESA = "Esquema Assessoria Contabil"
AVISO_LEGAL = "Extrato gerado para conferencia. Nao substitui o documento oficial do banco."
URL_ULTIMA_VERSAO = "https://api.github.com/repos/rodrigohsr/concilia/releases/latest"

# Paleta.
#
# Regra que orienta o uso de cor na tabela: cor marca excecao, nao a regra. Um
# extrato normal e quase todo debito, entao pintar cada linha de vermelho e
# verde faz a cor perder o sentido e cansa a leitura. O texto fica neutro e a
# cor aparece so onde ajuda a achar algo.
COR_FUNDO = "#f1f5f9"        # fundo da janela
COR_SUPERFICIE = "#ffffff"   # cartoes e tabela
COR_BORDA = "#e2e8f0"
COR_BORDA_FORTE = "#cbd5e1"
COR_TEXTO = "#0f172a"
COR_TEXTO_FRACO = "#64748b"
COR_DESTAQUE = "#0284c7"
COR_DESTAQUE_FUNDO = "#e0f2fe"
COR_DEBITO = "#b91c1c"
COR_CREDITO = "#15803d"
COR_ZEBRA = "#f8fafc"

FONTE = "Segoe UI"
FONTE_NUM = "Consolas"      # numeros alinham melhor em fonte de largura fixa

# Quantidade de linhas inseridas por ciclo ao preencher a tabela. Manter a
# insercao fatiada deixa a janela responsiva em extratos muito grandes.
LOTE_LINHAS = 400


def resource_path(relative_path: str) -> str:
    """Caminho de um recurso, funcionando tambem dentro do bundle do PyInstaller."""
    base = getattr(sys, "_MEIPASS", None) or os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base, relative_path)


def caminho_config() -> str:
    base = os.environ.get("APPDATA") or os.path.expanduser("~")
    return os.path.join(base, APP_NOME, "config.json")


def formatar_moeda(valor: float | None, vazio: str = "-") -> str:
    """Formata no padrao brasileiro: 1234.5 -> '1.234,50'."""
    if valor is None:
        return vazio
    return f"{valor:,.2f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def formatar_valor_cd(valor: float | None, vazio: str = "-") -> str:
    """Formata com sufixo contabil C (credito) ou D (debito)."""
    if valor is None:
        return vazio
    return f"{formatar_moeda(abs(valor))} {'C' if valor >= 0 else 'D'}"


def escapar_xml(texto: str) -> str:
    """Escapa &, < e > para uso dentro de Paragraph do reportlab.

    Historicos bancarios trazem '&' com frequencia ("JOAO & MARIA LTDA"), o que
    quebraria a geracao do PDF se fosse repassado cru.
    """
    return texto.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def formatar_data(iso: str | None) -> str:
    if not iso or len(iso) < 10:
        return ""
    return f"{iso[8:10]}/{iso[5:7]}/{iso[0:4]}"


@dataclass(slots=True)
class Linha:
    """Uma linha da tabela, com os campos ja formatados para exibicao."""

    iso: str
    data: str
    tipo: str
    historico: str          # o texto exibido: o do banco, ou o editado
    valor: float
    saldo: float | None
    valor_txt: str
    saldo_txt: str
    busca: str              # historico + tipo em minusculas, para filtro rapido
    chave: str              # identificador estavel do lancamento, para as edicoes
    original: str           # historico como veio do banco, para poder desfazer
    fit_id: str | None = None
    check_num: str | None = None
    ref_num: str | None = None

    @property
    def editado(self) -> bool:
        return self.historico != self.original


class ConciliaApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.config = self._carregar_config()

        self.extratos: list[ExtratoOFX] = []
        self.extrato: ExtratoOFX | None = None
        self.linhas: list[Linha] = []
        self.visiveis: list[Linha] = []
        self.caminho_atual: str | None = None
        self.ordenacao: tuple[str, bool] = ("data", False)
        self.geracao = 0  # invalida preenchimentos em andamento
        self._busca_agendada: str | None = None

        self.modo_edicao = False
        self.edicoes: dict[str, str] = {}   # chave do lancamento -> historico corrigido
        self.editor: tk.Entry | None = None
        self.editor_iid: str | None = None

        self._montar_janela()
        self._montar_cabecalho()
        self._montar_barra()
        self._montar_tabela()
        self._montar_rodape()
        self._montar_atalhos()

        # Tudo que nao e essencial para a janela aparecer fica para depois do
        # primeiro desenho: drag-and-drop, logo e o arquivo passado por linha
        # de comando.
        self.root.after_idle(self._pos_inicializacao)

    # ------------------------------------------------------------------
    # Construcao da interface
    # ------------------------------------------------------------------
    def _montar_janela(self) -> None:
        # a versao no titulo torna imediato saber qual build esta rodando
        self.root.title(f"{APP_NOME} {VERSAO} - {EMPRESA}")
        self.root.geometry(self.config.get("geometria", "1200x680"))
        self.root.minsize(900, 480)
        for icone in ("concilia.ico", "cifrao.ico"):
            try:
                self.root.iconbitmap(resource_path(icone))
                break
            except tk.TclError:
                continue

        self.root.configure(bg=COR_FUNDO)
        self._aplicar_tema()
        self.root.protocol("WM_DELETE_WINDOW", self._ao_fechar)

    def _aplicar_tema(self) -> None:
        """Estiliza o ttk. O 'clam' e a base porque e o tema que mais aceita
        customizacao - os temas nativos do Windows ignoram boa parte das cores."""
        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass

        e = self.style.configure
        m = self.style.map

        # --- tabela ---------------------------------------------------------
        e(
            "Concilia.Treeview",
            background=COR_SUPERFICIE,
            fieldbackground=COR_SUPERFICIE,
            foreground=COR_TEXTO,
            rowheight=30,
            borderwidth=0,
            relief="flat",
            font=(FONTE, 9),
        )
        m(
            "Concilia.Treeview",
            background=[("selected", COR_DESTAQUE_FUNDO)],
            foreground=[("selected", COR_TEXTO)],
        )
        e(
            "Concilia.Treeview.Heading",
            background=COR_FUNDO,
            foreground=COR_TEXTO_FRACO,
            relief="flat",
            borderwidth=0,
            padding=(12, 10),
            font=(FONTE, 8, "bold"),
        )
        m(
            "Concilia.Treeview.Heading",
            background=[("active", COR_BORDA)],
            foreground=[("active", COR_TEXTO)],
        )
        # remove a moldura afundada que o clam desenha em volta da tabela
        self.style.layout(
            "Concilia.Treeview",
            [("Concilia.Treeview.treearea", {"sticky": "nswe"})],
        )

        # --- botoes ---------------------------------------------------------
        for nome, fundo, texto, fundo_hover in (
            ("Acao.TButton", COR_DESTAQUE, "#ffffff", "#0369a1"),
            ("Secundario.TButton", COR_SUPERFICIE, COR_TEXTO, COR_DESTAQUE_FUNDO),
        ):
            e(
                nome,
                background=fundo,
                foreground=texto,
                bordercolor=COR_BORDA_FORTE,
                lightcolor=fundo,
                darkcolor=fundo,
                focuscolor=fundo,
                borderwidth=0 if nome == "Acao.TButton" else 1,
                relief="flat",
                padding=(14, 7),
                font=(FONTE, 9),
            )
            m(
                nome,
                background=[("disabled", COR_FUNDO), ("pressed", fundo_hover), ("active", fundo_hover)],
                foreground=[("disabled", "#94a3b8")],
                bordercolor=[("disabled", COR_BORDA)],
            )

        # --- campos ---------------------------------------------------------
        e(
            "Concilia.TEntry",
            fieldbackground=COR_SUPERFICIE,
            foreground=COR_TEXTO,
            bordercolor=COR_BORDA_FORTE,
            lightcolor=COR_BORDA_FORTE,
            darkcolor=COR_BORDA_FORTE,
            insertcolor=COR_TEXTO,
            borderwidth=1,
            relief="flat",
            padding=(8, 6),
        )
        m("Concilia.TEntry", bordercolor=[("focus", COR_DESTAQUE)], lightcolor=[("focus", COR_DESTAQUE)])

        e(
            "Concilia.TCombobox",
            fieldbackground=COR_SUPERFICIE,
            background=COR_SUPERFICIE,
            foreground=COR_TEXTO,
            bordercolor=COR_BORDA_FORTE,
            lightcolor=COR_BORDA_FORTE,
            darkcolor=COR_BORDA_FORTE,
            arrowcolor=COR_TEXTO_FRACO,
            borderwidth=1,
            padding=(8, 5),
        )
        m(
            "Concilia.TCombobox",
            fieldbackground=[("readonly", COR_SUPERFICIE)],
            bordercolor=[("focus", COR_DESTAQUE), ("active", COR_DESTAQUE)],
            arrowcolor=[("active", COR_DESTAQUE)],
        )
        # a lista suspensa nao e um widget ttk; so o Tcl alcanca as cores dela
        self.root.option_add("*TCombobox*Listbox.background", COR_SUPERFICIE)
        self.root.option_add("*TCombobox*Listbox.foreground", COR_TEXTO)
        self.root.option_add("*TCombobox*Listbox.selectBackground", COR_DESTAQUE_FUNDO)
        self.root.option_add("*TCombobox*Listbox.selectForeground", COR_TEXTO)
        self.root.option_add("*TCombobox*Listbox.font", (FONTE, 9))

        # --- barra de rolagem -----------------------------------------------
        e(
            "Concilia.Vertical.TScrollbar",
            background=COR_BORDA_FORTE,
            troughcolor=COR_SUPERFICIE,
            bordercolor=COR_SUPERFICIE,
            lightcolor=COR_SUPERFICIE,
            darkcolor=COR_SUPERFICIE,
            arrowcolor=COR_TEXTO_FRACO,
            borderwidth=0,
            width=12,
        )
        m("Concilia.Vertical.TScrollbar", background=[("active", COR_TEXTO_FRACO)])

    def _montar_cabecalho(self) -> None:
        self.header = tk.Frame(self.root, bg=COR_SUPERFICIE)
        self.header.pack(fill="x", side="top")

        identidade = tk.Frame(self.header, bg=COR_SUPERFICIE)
        identidade.pack(fill="x", padx=20, pady=(16, 4))

        # marca a esquerda; o logo da empresa, se houver, entra no lugar
        self.titulo = tk.Label(
            identidade,
            text=APP_NOME,
            font=(FONTE, 15, "bold"),
            fg=COR_DESTAQUE,
            bg=COR_SUPERFICIE,
        )
        self.titulo.pack(side="left", anchor="w")

        # Identidade da conta a direita, em duas alturas: o banco em destaque e
        # os detalhes (agencia, conta, periodo) abaixo, em tom fraco. Antes era
        # tudo uma linha corrida, com o mesmo peso para banco e agencia.
        bloco_conta = tk.Frame(identidade, bg=COR_SUPERFICIE)
        bloco_conta.pack(side="right", anchor="e")

        self.lbl_banco = tk.Label(
            bloco_conta,
            text="Nenhum extrato aberto",
            font=(FONTE, 11, "bold"),
            fg=COR_TEXTO,
            bg=COR_SUPERFICIE,
            anchor="e",
        )
        self.lbl_banco.pack(anchor="e")

        self.lbl_conta = tk.Label(
            bloco_conta,
            text="Abra um arquivo OFX ou arraste-o para esta janela",
            font=(FONTE, 9),
            fg=COR_TEXTO_FRACO,
            bg=COR_SUPERFICIE,
            anchor="e",
        )
        self.lbl_conta.pack(anchor="e")

        # --- cartoes de resumo ----------------------------------------------
        self.cartoes = tk.Frame(self.header, bg=COR_SUPERFICIE)
        self.cartoes.pack(fill="x", padx=20, pady=(12, 16))

        self.valores_cartao: dict[str, tk.Label] = {}
        definicao = (
            ("saldo_anterior", "Saldo anterior", COR_TEXTO),
            ("creditos", "Créditos", COR_CREDITO),
            ("debitos", "Débitos", COR_DEBITO),
            ("saldo_final", "Saldo final", COR_TEXTO),
        )
        for coluna, (chave, rotulo, cor) in enumerate(definicao):
            # destaque so no cartao do saldo final: e o numero que se confere
            final = chave == "saldo_final"
            fundo = COR_DESTAQUE_FUNDO if final else COR_FUNDO

            cartao = tk.Frame(
                self.cartoes,
                bg=fundo,
                highlightthickness=1,
                highlightbackground=COR_DESTAQUE if final else COR_BORDA,
            )
            cartao.grid(row=0, column=coluna, sticky="ew", padx=(0, 10) if coluna < 3 else 0)
            self.cartoes.columnconfigure(coluna, weight=1, uniform="cartao")

            tk.Label(
                cartao,
                text=rotulo.upper(),
                font=(FONTE, 8),
                fg=COR_TEXTO_FRACO,
                bg=fundo,
                anchor="w",
            ).pack(fill="x", padx=14, pady=(10, 0))

            valor = tk.Label(
                cartao,
                text="-",
                font=(FONTE, 15 if final else 14, "bold"),
                fg=cor,
                bg=fundo,
                anchor="w",
            )
            valor.pack(fill="x", padx=14, pady=(0, 10))
            self.valores_cartao[chave] = valor

        tk.Frame(self.root, bg=COR_BORDA, height=1).pack(fill="x")

    def _montar_barra(self) -> None:
        barra = tk.Frame(self.root, bg=COR_FUNDO)
        barra.pack(fill="x", padx=20, pady=(14, 10))

        self.btn_abrir = ttk.Button(
            barra, text="Abrir extrato", style="Acao.TButton", command=self.selecionar_arquivo
        )
        self.btn_abrir.pack(side="left")

        self.btn_recarregar = ttk.Button(
            barra, text="Recarregar", style="Secundario.TButton", command=self.recarregar, state="disabled"
        )
        self.btn_recarregar.pack(side="left", padx=(8, 0))

        self.btn_pdf = ttk.Button(
            barra, text="PDF", style="Secundario.TButton", command=self.exportar_pdf, state="disabled"
        )
        self.btn_pdf.pack(side="left", padx=(8, 0))

        self.btn_planilha = ttk.Button(
            barra, text="Planilha", style="Secundario.TButton", command=self.exportar_planilha, state="disabled"
        )
        self.btn_planilha.pack(side="left", padx=(8, 0))

        self.btn_editar = ttk.Button(
            barra, text="Editar", style="Secundario.TButton", command=self._alternar_modo_edicao, state="disabled"
        )
        self.btn_editar.pack(side="left", padx=(8, 0))

        # so aparece quando o extrato atual tem historicos corrigidos a mao
        self.btn_restaurar = ttk.Button(
            barra, text="Restaurar", style="Secundario.TButton", command=self._restaurar_historicos
        )

        # Seletor de conta: so aparece em arquivos com mais de um extrato
        self.frame_conta = tk.Frame(barra, bg=COR_FUNDO)
        self.cb_conta = ttk.Combobox(self.frame_conta, state="readonly", width=32, style="Concilia.TCombobox")
        self.cb_conta.pack(side="left")
        self.cb_conta.bind("<<ComboboxSelected>>", self._trocar_conta)

        # Busca e filtro ficam encostados na direita
        self.var_busca = tk.StringVar()
        self.var_filtro = tk.StringVar(value="Todos")

        self.cb_filtro = ttk.Combobox(
            barra,
            state="readonly",
            width=11,
            values=("Todos", "Créditos", "Débitos"),
            textvariable=self.var_filtro,
            style="Concilia.TCombobox",
        )
        self.cb_filtro.pack(side="right")
        self.cb_filtro.bind("<<ComboboxSelected>>", lambda _e: self._aplicar_filtros())

        self.entry_busca = ttk.Entry(barra, textvariable=self.var_busca, width=32, style="Concilia.TEntry")
        self.entry_busca.pack(side="right", padx=(0, 8))
        self.var_busca.trace_add("write", self._busca_alterada)
        self._marca_dagua_busca()

    def _marca_dagua_busca(self) -> None:
        """Texto de dica dentro do campo de busca.

        E um rotulo posicionado sobre o campo, e nao um texto pre-preenchido:
        assim a dica nunca chega a `var_busca` e nao corre o risco de ser
        confundida com um termo de pesquisa.
        """
        self.marca_busca = tk.Label(
            self.entry_busca, text="Localizar lançamento", font=(FONTE, 9), fg="#94a3b8", bg=COR_SUPERFICIE
        )
        self.marca_busca.bind("<Button-1>", lambda _e: self.entry_busca.focus_set())

        def alternar(*_args) -> None:
            if self.var_busca.get():
                self.marca_busca.place_forget()
            else:
                self.marca_busca.place(x=10, rely=0.5, anchor="w")

        self.var_busca.trace_add("write", alternar)
        alternar()

    def _montar_tabela(self) -> None:
        moldura = tk.Frame(self.root, bg=COR_SUPERFICIE, highlightthickness=1, highlightbackground=COR_BORDA)
        moldura.pack(fill="both", expand=True, padx=20, pady=(0, 12))

        colunas = ("data", "tipo", "historico", "valor", "saldo")
        self.tree = ttk.Treeview(
            moldura, columns=colunas, show="headings", selectmode="extended", style="Concilia.Treeview"
        )

        titulos = {
            "data": ("Data", 96, "center", False),
            "tipo": ("Tipo", 96, "w", False),
            "historico": ("Histórico", 520, "center", True),
            "valor": ("Valor", 140, "e", False),
            "saldo": ("Saldo", 140, "e", False),
        }
        for coluna, (texto, largura, alinhamento, estica) in titulos.items():
            self.tree.heading(coluna, text=texto, command=lambda c=coluna: self._ordenar_por(c))
            self.tree.column(coluna, width=largura, anchor=alinhamento, stretch=estica, minwidth=70)

        # Cor so no debito, e so no tom do texto: e a excecao que se procura num
        # extrato. Credito fica neutro - antes as duas cores brigavam em todas
        # as linhas e nenhuma delas informava nada.
        self.tree.tag_configure("zebra", background=COR_ZEBRA)
        # historico corrigido a mao e excecao, entao aqui a cor informa
        self.tree.tag_configure("editado", foreground=COR_DESTAQUE)

        vsb = ttk.Scrollbar(moldura, orient="vertical", command=self.tree.yview, style="Concilia.Vertical.TScrollbar")
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(1, 0), pady=1)
        vsb.grid(row=0, column=1, sticky="ns", pady=1, padx=(0, 1))
        moldura.rowconfigure(0, weight=1)
        moldura.columnconfigure(0, weight=1)

        self.tree.bind("<Double-1>", self._duplo_clique)

    def _montar_rodape(self) -> None:
        rodape = tk.Frame(self.root, bg=COR_FUNDO)
        rodape.pack(fill="x", side="bottom")
        self.status = tk.Label(
            rodape,
            text="Pronto",
            bg=COR_FUNDO,
            fg=COR_TEXTO_FRACO,
            anchor="w",
            font=(FONTE, 9),
            padx=20,
            pady=8,
        )
        self.status.pack(side="left")

        # so aparece quando houver versao nova publicada
        self.aviso_atualizacao = tk.Label(
            rodape,
            text="",
            bg=COR_FUNDO,
            fg=COR_DESTAQUE,
            font=(FONTE, 9, "bold"),
            cursor="hand2",
            pady=8,
        )
        self.aviso_atualizacao.bind("<Button-1>", self._abrir_pagina_atualizacao)

    def _montar_atalhos(self) -> None:
        self.root.bind("<Control-o>", lambda _e: self.selecionar_arquivo())
        self.root.bind("<F5>", lambda _e: self.recarregar())
        self.root.bind("<Control-f>", lambda _e: self.entry_busca.focus_set())
        self.root.bind("<Escape>", self._limpar_busca)
        self.root.bind("<Control-p>", lambda _e: self.exportar_pdf())
        self.tree.bind("<Control-c>", self._copiar_historico)
        self.tree.bind("<Control-C>", self._copiar_linhas)  # Ctrl+Shift+C
        self.tree.bind("<Control-a>", self._selecionar_tudo)
        self.root.bind("<F2>", self._editar_selecionado)
        self.root.bind("<Control-e>", lambda _e: self._alternar_modo_edicao())

    def _pos_inicializacao(self) -> None:
        self._carregar_logo()
        self._ativar_arrastar_soltar()
        if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
            self.carregar(sys.argv[1])
        self._verificar_atualizacao()

    # ------------------------------------------------------------------
    # Verificacao de atualizacao
    # ------------------------------------------------------------------
    def _verificar_atualizacao(self) -> None:
        """Consulta a ultima versao publicada no GitHub, sem atrasar nada.

        A consulta roda numa thread e o resultado e recolhido pelo laco do
        Tkinter: widget de interface so pode ser tocado pela thread principal.
        Falha de rede e silenciosa - o programa funciona offline.
        """
        if not self.config.get("verificar_atualizacoes", True):
            return

        self._versao_publicada: list = []

        def consultar() -> None:
            try:
                import json as _json
                import urllib.request

                requisicao = urllib.request.Request(
                    URL_ULTIMA_VERSAO,
                    headers={"Accept": "application/vnd.github+json", "User-Agent": f"{APP_NOME}/{VERSAO}"},
                )
                with urllib.request.urlopen(requisicao, timeout=8) as resposta:
                    dados = _json.load(resposta)
                self._versao_publicada.append((dados.get("tag_name") or "", dados.get("html_url") or ""))
            except Exception:
                self._versao_publicada.append(None)

        import threading

        threading.Thread(target=consultar, daemon=True).start()
        self.root.after(1200, self._recolher_resultado_atualizacao, 0)

    def _recolher_resultado_atualizacao(self, tentativa: int) -> None:
        if not self._versao_publicada:
            if tentativa < 12:  # ~12 s de espera, depois desiste em silencio
                self.root.after(1000, self._recolher_resultado_atualizacao, tentativa + 1)
            return

        resultado = self._versao_publicada[0]
        if not resultado:
            return
        tag, endereco = resultado
        publicada = self._numero_versao(tag)
        if not publicada or publicada <= self._numero_versao(VERSAO):
            return

        self.endereco_atualizacao = endereco
        self.aviso_atualizacao.configure(text=f"Versão {tag.lstrip('v')} disponível  ›")
        self.aviso_atualizacao.pack(side="right", padx=20)

    @staticmethod
    def _numero_versao(texto: str) -> tuple[int, ...]:
        """'v1.2.3' -> (1, 2, 3), para comparar versoes numericamente."""
        import re

        numeros = re.findall(r"\d+", texto or "")
        return tuple(int(n) for n in numeros[:3])

    def _abrir_pagina_atualizacao(self, _evento=None) -> None:
        import webbrowser

        if getattr(self, "endereco_atualizacao", ""):
            webbrowser.open(self.endereco_atualizacao)

    def _carregar_logo(self) -> None:
        """Usa o PhotoImage nativo do Tk (le PNG sem depender do Pillow)."""
        caminho = resource_path("logo.png")
        if not os.path.exists(caminho):
            return  # sem logo da empresa, o cabecalho mostra o nome do programa
        try:
            imagem = tk.PhotoImage(file=caminho)
            fator = max(1, round(imagem.height() / 65))
            if fator > 1:
                imagem = imagem.subsample(fator, fator)
        except tk.TclError:
            try:  # formato que o Tk nao le (JPEG, PNG exotico): tenta o Pillow
                from PIL import Image, ImageTk

                img = Image.open(caminho)
                largura = int(img.width * (65 / img.height))
                imagem = ImageTk.PhotoImage(img.resize((largura, 65), Image.LANCZOS))
            except Exception:
                return
        self.logo = imagem  # a referencia precisa sobreviver ao metodo
        self.titulo.configure(image=imagem, text="")

    def _ativar_arrastar_soltar(self) -> None:
        try:
            import windnd

            windnd.hook_dropfiles(self.root, func=self._ao_soltar_arquivo)
        except Exception:
            pass  # recurso opcional

    def _ao_soltar_arquivo(self, arquivos) -> None:
        if not arquivos:
            return
        caminho = arquivos[0].decode("mbcs" if os.name == "nt" else "utf-8", errors="replace")
        self.carregar(caminho)

    # ------------------------------------------------------------------
    # Leitura do arquivo
    # ------------------------------------------------------------------
    def selecionar_arquivo(self) -> None:
        caminho = filedialog.askopenfilename(
            title="Selecione o extrato",
            initialdir=self.config.get("ultima_pasta") or os.path.expanduser("~"),
            filetypes=[("Arquivos OFX", "*.ofx *.qfx *.OFX *.QFX"), ("Todos os arquivos", "*.*")],
        )
        if caminho:
            self.carregar(caminho)

    def recarregar(self) -> None:
        if self.caminho_atual:
            self.carregar(self.caminho_atual)

    def carregar(self, caminho: str) -> None:
        self.status.configure(text=f"Lendo {os.path.basename(caminho)}...")
        self.root.configure(cursor="watch")
        self.root.update_idletasks()
        try:
            extratos = OFXParserBR().parse_file_todos(caminho)
        except FileNotFoundError:
            self._erro("Arquivo nao encontrado", f"O arquivo nao existe mais:\n{caminho}")
            return
        except PermissionError:
            self._erro("Sem permissao", "O arquivo esta aberto em outro programa ou sem permissao de leitura.")
            return
        except Exception as exc:  # arquivo corrompido, formato inesperado...
            self._erro("Erro de leitura", f"Nao foi possivel ler este arquivo OFX:\n\n{exc}")
            return
        finally:
            self.root.configure(cursor="")

        if not extratos:
            self._erro(
                "Formato nao reconhecido",
                "Nenhum extrato foi encontrado no arquivo.\n\n"
                "Confirme se ele e realmente um OFX exportado pelo banco.",
            )
            return

        self._fechar_editor()
        self.caminho_atual = caminho
        self.extratos = extratos
        self.config["ultima_pasta"] = os.path.dirname(caminho)
        self._carregar_edicoes(caminho)

        if len(extratos) > 1:
            self.cb_conta.configure(values=[self._rotulo_conta(e) for e in extratos])
            self.cb_conta.current(0)
            self.frame_conta.pack(side="left", padx=4, pady=6)
        else:
            self.frame_conta.pack_forget()

        self.btn_recarregar.configure(state="normal")
        self._selecionar_extrato(0)

    def _rotulo_conta(self, extrato: ExtratoOFX) -> str:
        conta = extrato.conta
        tipo = conta.tipo_descricao or "Conta"
        banco = conta.banco_nome or conta.bank_id
        return f"{tipo} {conta.acct_id or '?'}" + (f" - {banco}" if banco else "")

    def _trocar_conta(self, _evento=None) -> None:
        self._selecionar_extrato(self.cb_conta.current())

    def _selecionar_extrato(self, indice: int) -> None:
        if not (0 <= indice < len(self.extratos)):
            return
        self.extrato = self.extratos[indice]
        self.linhas = [self._montar_linha(t) for t in self.extrato.transacoes]
        self.ordenacao = ("data", False)
        self._atualizar_cabecalho()
        self._aplicar_filtros()

    def _chave_lancamento(self, transacao) -> str:
        """Identificador estavel de um lancamento, para reencontrar sua edicao.

        O FITID e o identificador que o proprio banco garante unico. Quando ele
        falta, monta-se uma chave com data, valor e historico - o suficiente
        para nao confundir lancamentos dentro de um mesmo extrato.
        """
        conta = (self.extrato.conta.acct_id if self.extrato else None) or "?"
        if transacao.fit_id:
            return f"{conta}:{transacao.fit_id}"
        bruto = f"{transacao.dt_posted_iso}|{transacao.valor}|{(transacao.memo or transacao.name or '')[:60]}"
        return f"{conta}:~{bruto}"

    def _montar_linha(self, transacao) -> Linha:
        valor = transacao.valor or 0.0
        original = transacao.descricao or "(sem histórico)"
        chave = self._chave_lancamento(transacao)
        historico = self.edicoes.get(chave, original)
        tipo = transacao.tipo or ""
        return Linha(
            iso=transacao.dt_posted_iso or "",
            data=formatar_data(transacao.dt_posted_iso),
            tipo=tipo,
            historico=historico,
            valor=valor,
            saldo=transacao.saldo,
            valor_txt=formatar_valor_cd(valor),
            saldo_txt=formatar_valor_cd(transacao.saldo),
            busca=f"{historico}\n{tipo}".lower(),
            chave=chave,
            original=original,
            fit_id=transacao.fit_id,
            check_num=transacao.check_num,
            ref_num=transacao.ref_num,
        )

    @staticmethod
    def _detalhes_linha(linha: Linha) -> str:
        campos = [
            ("Data", linha.data),
            ("Tipo", linha.tipo),
            ("Valor", linha.valor_txt),
            ("Saldo após o lançamento", linha.saldo_txt),
            ("Documento", linha.check_num),
            ("Referência", linha.ref_num),
            ("Identificador (FITID)", linha.fit_id),
            ("Histórico", linha.historico),
        ]
        if linha.editado:
            campos.append(("Histórico original do banco", linha.original))
        return "\n".join(f"{rotulo}: {texto}" for rotulo, texto in campos if texto)

    def _detalhes_conta(self) -> str:
        """Linha secundaria do cabecalho: agencia, conta, periodo e contagem."""
        extrato = self.extrato
        if extrato is None:
            return ""
        conta = extrato.conta
        partes = []
        if conta.branch_id:
            partes.append(f"Ag. {conta.branch_id}")
        rotulo = "Cartão" if (conta.acct_type or "").upper() == "CREDITCARD" else "Conta"
        partes.append(f"{rotulo} {conta.acct_id or 'não informada'}")
        if extrato.dt_inicio_iso and extrato.dt_fim_iso:
            partes.append(f"{formatar_data(extrato.dt_inicio_iso)} a {formatar_data(extrato.dt_fim_iso)}")
        partes.append(f"{len(extrato.transacoes)} lançamentos")
        return "   ·   ".join(partes)

    def _atualizar_cabecalho(self) -> None:
        extrato = self.extrato
        if extrato is None:
            return

        self.lbl_banco.configure(text=extrato.conta.descricao)
        self.lbl_conta.configure(text=self._detalhes_conta())

        self._atualizar_cartoes()

        if extrato.saldo_final is not None:
            self.valores_cartao["saldo_anterior"].configure(
                text=f"R$ {formatar_moeda(extrato.saldo_inicial)}", fg=self._cor_saldo(extrato.saldo_inicial)
            )
            self.valores_cartao["saldo_final"].configure(
                text=f"R$ {formatar_moeda(extrato.saldo_final)}", fg=self._cor_saldo(extrato.saldo_final)
            )
        else:
            # o banco nao enviou LEDGERBAL: sem ele nao da para reconstruir a serie
            for chave in ("saldo_anterior", "saldo_final"):
                self.valores_cartao[chave].configure(text="não informado", fg=COR_TEXTO_FRACO)

    def _atualizar_cartoes(self) -> None:
        """Creditos e debitos seguem o que esta em tela; os saldos, nao.

        Saldo anterior e final vem do arquivo e valem para o periodo inteiro -
        filtrar a lista nao muda o saldo que o banco fechou.
        """
        creditos = sum(l.valor for l in self.visiveis if l.valor >= 0)
        debitos = sum(l.valor for l in self.visiveis if l.valor < 0)
        self.valores_cartao["creditos"].configure(text=f"R$ {formatar_moeda(creditos)}")
        self.valores_cartao["debitos"].configure(text=f"R$ {formatar_moeda(abs(debitos))}")

    @staticmethod
    def _cor_saldo(valor: float | None) -> str:
        return COR_DEBITO if (valor or 0.0) < 0 else COR_TEXTO

    # ------------------------------------------------------------------
    # Filtro, ordenacao e preenchimento da tabela
    # ------------------------------------------------------------------
    def _busca_alterada(self, *_args) -> None:
        # espera o usuario parar de digitar antes de refiltrar
        if self._busca_agendada is not None:
            self.root.after_cancel(self._busca_agendada)
        self._busca_agendada = self.root.after(180, self._aplicar_filtros)

    def _limpar_busca(self, _evento=None) -> None:
        if self.var_busca.get():
            self.var_busca.set("")

    def _aplicar_filtros(self) -> None:
        self._busca_agendada = None
        termo = self.var_busca.get().strip().lower()
        filtro = self.var_filtro.get()

        linhas = self.linhas
        if filtro == "Créditos":
            linhas = [l for l in linhas if l.valor >= 0]
        elif filtro == "Débitos":
            linhas = [l for l in linhas if l.valor < 0]
        if termo:
            termos = termo.split()
            linhas = [l for l in linhas if all(t in l.busca for t in termos)]

        self.visiveis = linhas
        self._atualizar_cartoes()
        self._ordenar_visiveis()
        self._preencher_tabela()
        self._atualizar_status()

    _CHAVES = {
        "data": lambda l: l.iso,
        "tipo": lambda l: l.tipo.lower(),
        "historico": lambda l: l.historico.lower(),
        "valor": lambda l: l.valor,
        "saldo": lambda l: (l.saldo is None, l.saldo or 0.0),
    }

    def _ordenar_por(self, coluna: str) -> None:
        atual, invertido = self.ordenacao
        self.ordenacao = (coluna, not invertido if coluna == atual else False)
        self._ordenar_visiveis()
        self._preencher_tabela()

    def _ordenar_visiveis(self) -> None:
        coluna, invertido = self.ordenacao
        self.visiveis.sort(key=self._CHAVES[coluna], reverse=invertido)
        for nome in self._CHAVES:
            texto = {"data": "Data", "tipo": "Tipo", "historico": "Histórico", "valor": "Valor", "saldo": "Saldo"}[nome]
            if nome == coluna:
                texto += "  ▼" if invertido else "  ▲"
            self.tree.heading(nome, text=texto)

    def _preencher_tabela(self) -> None:
        self.geracao += 1
        self.tree.delete(*self.tree.get_children())
        self._inserir_lote(0, self.geracao)

    def _inserir_lote(self, inicio: int, geracao: int) -> None:
        if geracao != self.geracao:
            return  # outro preenchimento comecou; este ficou obsoleto
        inserir = self.tree.insert
        fim = min(inicio + LOTE_LINHAS, len(self.visiveis))
        for i in range(inicio, fim):
            linha = self.visiveis[i]
            tags = ["editado"] if linha.editado else []
            if i % 2:
                tags.append("zebra")
            inserir(
                "",
                "end",
                iid=str(i),
                values=(linha.data, linha.tipo, linha.historico, linha.valor_txt, linha.saldo_txt),
                tags=tags,
            )
        if fim < len(self.visiveis):
            self.root.after(1, self._inserir_lote, fim, geracao)

    def _atualizar_status(self) -> None:
        # so faz sentido exportar o que esta visivel
        estado = "normal" if self.visiveis else "disabled"
        self.btn_pdf.configure(state=estado)
        self.btn_planilha.configure(state=estado)
        self.btn_editar.configure(state="normal" if self.linhas else "disabled")
        self._atualizar_botao_restaurar()

        if not self.linhas:
            self.status.configure(text="Nenhum lançamento neste extrato.")
            return
        if not self.visiveis:
            self.status.configure(
                text=f"Nenhum dos {len(self.linhas)} lançamentos corresponde ao filtro atual."
            )
            return

        contagem = (
            f"Exibindo {len(self.visiveis)} de {len(self.linhas)} lançamentos"
            if len(self.visiveis) != len(self.linhas)
            else f"{len(self.linhas)} lançamentos"
        )
        # Com filtro ativo os cartoes do topo passam a mostrar os totais do que
        # esta em tela, e nao do extrato inteiro; o aviso evita a leitura errada.
        if len(self.visiveis) != len(self.linhas):
            contagem += "   ·   totais referentes ao filtro"
        self.status.configure(text=contagem)

    # ------------------------------------------------------------------
    # Interacoes com a tabela
    # ------------------------------------------------------------------
    def _selecionar_tudo(self, _evento=None) -> str:
        self.tree.selection_set(self.tree.get_children())
        return "break"

    def _linhas_selecionadas(self) -> list[Linha]:
        return [self.visiveis[int(iid)] for iid in self.tree.selection() if iid.isdigit()]

    def _copiar_historico(self, _evento=None) -> str:
        """Ctrl+C: copia so o historico (o campo que costuma ir para o lancamento)."""
        linhas = self._linhas_selecionadas()
        if linhas:
            self._para_area_transferencia("\n".join(l.historico for l in linhas))
            self.status.configure(text=f"{len(linhas)} histórico(s) copiado(s).")
        return "break"

    def _copiar_linhas(self, _evento=None) -> str:
        """Ctrl+Shift+C: copia a linha inteira, pronta para colar no Excel."""
        linhas = self._linhas_selecionadas()
        if linhas:
            texto = "\n".join(
                "\t".join((l.data, l.tipo, l.historico, formatar_moeda(l.valor), formatar_moeda(l.saldo, "")))
                for l in linhas
            )
            self._para_area_transferencia(texto)
            self.status.configure(text=f"{len(linhas)} linha(s) copiada(s) para colar em planilha.")
        return "break"

    def _para_area_transferencia(self, texto: str) -> None:
        self.root.clipboard_clear()
        self.root.clipboard_append(texto)
        self.root.update_idletasks()

    def _mostrar_detalhes(self, _evento=None) -> None:
        linhas = self._linhas_selecionadas()
        if linhas:
            messagebox.showinfo("Detalhes do lançamento", self._detalhes_linha(linhas[0]), parent=self.root)

    # ------------------------------------------------------------------
    # Modo de edicao do historico
    # ------------------------------------------------------------------
    def _caminho_edicoes(self, caminho_ofx: str) -> str:
        """Arquivo de edicoes correspondente a um OFX.

        Fica em %APPDATA%, e nao ao lado do extrato: o arquivo do cliente e
        so lido, nunca modificado, e a pasta de origem pode ser somente leitura
        ou estar numa rede. O nome vem do caminho completo, para dois extratos
        de mesmo nome em pastas diferentes nao se misturarem.
        """
        import hashlib

        digest = hashlib.sha1(os.path.abspath(caminho_ofx).lower().encode("utf-8")).hexdigest()[:16]
        return os.path.join(os.path.dirname(caminho_config()), "edicoes", f"{digest}.json")

    def _carregar_edicoes(self, caminho_ofx: str) -> None:
        self.edicoes = {}
        try:
            with open(self._caminho_edicoes(caminho_ofx), encoding="utf-8") as arquivo:
                dados = json.load(arquivo)
            historicos = dados.get("historicos")
            if isinstance(historicos, dict):
                self.edicoes = {c: t for c, t in historicos.items() if isinstance(t, str)}
        except FileNotFoundError:
            pass
        except Exception:
            pass  # arquivo corrompido nao pode impedir a abertura do extrato

    def _salvar_edicoes(self) -> None:
        if not self.caminho_atual:
            return
        destino = self._caminho_edicoes(self.caminho_atual)
        try:
            if not self.edicoes:
                if os.path.exists(destino):
                    os.remove(destino)
                return
            os.makedirs(os.path.dirname(destino), exist_ok=True)
            with open(destino, "w", encoding="utf-8") as arquivo:
                json.dump(
                    {"versao": 1, "extrato": self.caminho_atual, "historicos": self.edicoes},
                    arquivo,
                    ensure_ascii=False,
                    indent=2,
                )
        except Exception as exc:
            self.status.configure(text=f"Não foi possível gravar as edições: {exc}")

    def _alternar_modo_edicao(self) -> None:
        self.modo_edicao = not self.modo_edicao
        self.btn_editar.configure(
            text="Editando" if self.modo_edicao else "Editar",
            style="Acao.TButton" if self.modo_edicao else "Secundario.TButton",
        )
        if self.modo_edicao:
            self.status.configure(
                text="Modo de edição: duplo clique ou F2 sobre o lançamento para corrigir o histórico."
                "   ·   Enter confirma, Esc cancela, campo vazio restaura o texto do banco."
            )
        else:
            self._fechar_editor()
            self._atualizar_status()

    def _duplo_clique(self, evento) -> str | None:
        """Em modo de edicao abre o editor; fora dele, mostra os detalhes."""
        if not self.modo_edicao:
            self._mostrar_detalhes()
            return None
        iid = self.tree.identify_row(evento.y)
        if iid:
            self._abrir_editor(iid)
        return "break"

    def _editar_selecionado(self, _evento=None) -> str:
        if self.modo_edicao:
            selecao = self.tree.selection()
            if selecao:
                self._abrir_editor(selecao[0])
        return "break"

    def _abrir_editor(self, iid: str) -> None:
        if not iid.isdigit() or int(iid) >= len(self.visiveis):
            return
        self._fechar_editor()

        self.tree.see(iid)
        caixa = self.tree.bbox(iid, "historico")
        if not caixa:  # linha fora da area visivel
            return
        x, y, largura, altura = caixa

        linha = self.visiveis[int(iid)]
        self.editor_iid = iid
        self.editor = tk.Entry(
            self.tree,
            font=(FONTE, 9),
            justify="center",
            relief="solid",
            borderwidth=1,
            highlightthickness=1,
            highlightcolor=COR_DESTAQUE,
            highlightbackground=COR_DESTAQUE,
        )
        self.editor.insert(0, linha.historico)
        self.editor.select_range(0, "end")
        self.editor.place(x=x, y=y, width=largura, height=altura)
        self.editor.focus_set()

        self.editor.bind("<Return>", lambda _e: self._confirmar_edicao())
        self.editor.bind("<KP_Enter>", lambda _e: self._confirmar_edicao())
        self.editor.bind("<Escape>", lambda _e: self._fechar_editor())
        self.editor.bind("<FocusOut>", lambda _e: self._confirmar_edicao())

    def _fechar_editor(self) -> None:
        if self.editor is not None:
            editor, self.editor = self.editor, None
            self.editor_iid = None
            editor.destroy()

    def _confirmar_edicao(self) -> None:
        if self.editor is None or self.editor_iid is None:
            return
        iid, texto = self.editor_iid, self.editor.get().strip()
        self._fechar_editor()

        if not iid.isdigit() or int(iid) >= len(self.visiveis):
            return
        linha = self.visiveis[int(iid)]

        # campo vazio devolve o texto original do banco
        novo = texto or linha.original
        if novo == linha.historico:
            return

        linha.historico = novo
        linha.busca = f"{novo}\n{linha.tipo}".lower()
        if novo == linha.original:
            self.edicoes.pop(linha.chave, None)
        else:
            self.edicoes[linha.chave] = novo

        self.tree.item(iid, values=(linha.data, linha.tipo, novo, linha.valor_txt, linha.saldo_txt))
        self._marcar_linha(iid, int(iid), linha)
        self._salvar_edicoes()
        self._atualizar_botao_restaurar()

        editados = sum(1 for l in self.linhas if l.editado)
        self.status.configure(
            text=f"Histórico atualizado.   ·   {editados} lançamento(s) editado(s) neste extrato."
        )

    def _atualizar_botao_restaurar(self) -> None:
        """Botao de desfazer so existe enquanto ha o que desfazer."""
        if self.edicoes:
            self.btn_restaurar.pack(side="left", padx=(8, 0), after=self.btn_editar)
        else:
            self.btn_restaurar.pack_forget()

    def _marcar_linha(self, iid: str, indice: int, linha: Linha) -> None:
        """Aplica as tags visuais de uma linha (zebra e marca de editado)."""
        tags = ["editado"] if linha.editado else []
        if indice % 2:
            tags.append("zebra")
        self.tree.item(iid, tags=tags)

    def _restaurar_historicos(self) -> None:
        if not self.edicoes:
            return
        quantos = len(self.edicoes)
        if not messagebox.askyesno(
            "Restaurar históricos",
            f"Desfazer {quantos} edição(ões) e voltar ao texto original do banco?",
            parent=self.root,
        ):
            return
        self.edicoes = {}
        self._salvar_edicoes()
        for linha in self.linhas:
            linha.historico = linha.original
            linha.busca = f"{linha.original}\n{linha.tipo}".lower()
        self._fechar_editor()
        self._aplicar_filtros()
        self.status.configure(text=f"{quantos} edição(ões) desfeita(s).")

    # ------------------------------------------------------------------
    # Exportacoes
    # ------------------------------------------------------------------
    def _nome_sugerido(self, extensao: str) -> str:
        base = os.path.splitext(os.path.basename(self.caminho_atual or "extrato"))[0]
        conta = (self.extrato.conta.acct_id if self.extrato else "") or ""
        conta = "".join(c for c in conta if c.isalnum())
        return f"{base}{'_' + conta if conta else ''}{extensao}"

    def exportar_pdf(self) -> None:
        if not self.visiveis:
            return
        destino = filedialog.asksaveasfilename(
            title="Salvar PDF",
            defaultextension=".pdf",
            initialfile=self._nome_sugerido(".pdf"),
            initialdir=self.config.get("ultima_pasta"),
            filetypes=[("Documento PDF", "*.pdf")],
        )
        if not destino:
            return
        try:
            self._gerar_pdf(destino)
        except ImportError:
            self._erro(
                "Biblioteca ausente",
                "A geracao de PDF depende do pacote reportlab.\n\nInstale com:\n    pip install reportlab",
            )
            return
        except PermissionError:
            self._erro("Arquivo em uso", "Feche o PDF no leitor antes de gravar por cima.")
            return
        except Exception as exc:
            self._erro("Erro ao salvar", f"Nao foi possivel gerar o PDF:\n\n{exc}")
            return
        self._sucesso(destino, "PDF gerado com sucesso.")

    def _gerar_pdf(self, destino: str) -> None:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle

        estilo_titulo = ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=14, alignment=TA_CENTER, leading=18)
        estilo_info = ParagraphStyle("info", fontName="Helvetica", fontSize=8, alignment=TA_CENTER, leading=11)
        estilo_celula = ParagraphStyle("celula", fontName="Helvetica", fontSize=7, leading=8.5)
        estilo_aviso = ParagraphStyle("aviso", fontName="Helvetica-Oblique", fontSize=7, alignment=TA_CENTER)

        doc = SimpleDocTemplate(
            destino,
            pagesize=A4,
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=12 * mm,
            bottomMargin=14 * mm,
            title=f"Extrato - {self.extrato.conta.acct_id if self.extrato else ''}",
            author=EMPRESA,
        )

        cabecalho = escapar_xml(self._resumo_para_pdf()).replace("\n", "<br/>")
        elementos = [
            Paragraph("Extrato Bancario", estilo_titulo),
            Spacer(1, 4),
            Paragraph(cabecalho, estilo_info),
            Spacer(1, 8),
        ]
        if len(self.visiveis) != len(self.linhas):
            elementos.append(Paragraph("<b>Atencao:</b> exportacao filtrada - nao contem todos os lancamentos.", estilo_info))
            elementos.append(Spacer(1, 6))

        dados = [["Data", "Tipo", "Historico", "Valor", "Saldo"]]
        for linha in self.visiveis:
            # o historico vai como Paragraph para quebrar em varias linhas em vez
            # de ser cortado; precisa ter &, < e > escapados
            dados.append(
                [
                    linha.data,
                    linha.tipo,
                    Paragraph(escapar_xml(linha.historico), estilo_celula),
                    linha.valor_txt,
                    linha.saldo_txt,
                ]
            )

        creditos = sum(l.valor for l in self.visiveis if l.valor >= 0)
        debitos = sum(l.valor for l in self.visiveis if l.valor < 0)
        dados.append(["", "", "Total de creditos", formatar_moeda(creditos), ""])
        dados.append(["", "", "Total de debitos", formatar_moeda(abs(debitos)), ""])
        dados.append(["", "", "Resultado do periodo", formatar_valor_cd(creditos + debitos), ""])

        tabela = LongTable(
            dados,
            colWidths=[20 * mm, 24 * mm, 82 * mm, 32 * mm, 32 * mm],
            repeatRows=1,
        )
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eef5")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9aa5b1")),
            ("ALIGN", (0, 0), (1, -1), "CENTER"),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 1), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
            ("FONTNAME", (0, -3), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -3), (-1, -1), colors.HexColor("#e8eef5")),
            ("ALIGN", (2, -3), (2, -1), "RIGHT"),
        ]
        for indice, linha in enumerate(self.visiveis, start=1):
            if linha.valor < 0:
                estilo.append(("TEXTCOLOR", (3, indice), (3, indice), colors.HexColor(COR_DEBITO)))
            if indice % 2 == 0:
                estilo.append(("BACKGROUND", (0, indice), (-1, indice), colors.HexColor(COR_ZEBRA)))
        tabela.setStyle(TableStyle(estilo))

        elementos.append(tabela)
        elementos.append(Spacer(1, 8))
        elementos.append(Paragraph(AVISO_LEGAL, estilo_aviso))

        def rodape(canvas, documento):
            canvas.saveState()
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#666666"))
            canvas.drawString(10 * mm, 8 * mm, f"{EMPRESA} - {os.path.basename(self.caminho_atual or '')}")
            canvas.drawRightString(A4[0] - 10 * mm, 8 * mm, f"Pagina {documento.page}")
            canvas.restoreState()

        doc.build(elementos, onFirstPage=rodape, onLaterPages=rodape)

    def _resumo_para_pdf(self) -> str:
        """Cabecalho do PDF, montado a partir do extrato.

        Nao le mais o texto da tela: os dados agora estao distribuidos entre
        varios rotulos e cartoes, e o PDF precisa da sua propria versao.
        """
        extrato = self.extrato
        if extrato is None:
            return ""
        linhas = [f"{extrato.conta.descricao}   |   {self._detalhes_conta()}"]
        if extrato.saldo_final is not None:
            linhas.append(
                f"Saldo anterior: R$ {formatar_valor_cd(extrato.saldo_inicial)}"
                f"   |   Saldo final: R$ {formatar_valor_cd(extrato.saldo_final)}"
            )
        else:
            linhas.append("Saldos nao informados no arquivo pelo banco")
        return "\n".join(linhas)

    def exportar_planilha(self) -> None:
        if not self.visiveis:
            return
        destino = filedialog.asksaveasfilename(
            title="Salvar planilha",
            defaultextension=".xlsx",
            initialfile=self._nome_sugerido(".xlsx"),
            initialdir=self.config.get("ultima_pasta"),
            filetypes=[("Planilha do Excel", "*.xlsx"), ("CSV para Excel", "*.csv")],
        )
        if not destino:
            return
        try:
            if destino.lower().endswith(".csv"):
                self._gerar_csv(destino)
            else:
                self._gerar_xlsx(destino)
        except ImportError:
            self._erro(
                "Biblioteca ausente",
                "O arquivo .xlsx depende do pacote openpyxl.\n\n"
                "Instale com:\n    pip install openpyxl\n\nOu salve como .csv, que nao exige nada.",
            )
            return
        except PermissionError:
            self._erro("Arquivo em uso", "Feche a planilha no Excel antes de gravar por cima.")
            return
        except Exception as exc:
            self._erro("Erro ao salvar", f"Nao foi possivel gerar a planilha:\n\n{exc}")
            return
        self._sucesso(destino, "Planilha gerada com sucesso.")

    def _gerar_csv(self, destino: str) -> None:
        import csv

        # utf-8-sig + ';' fazem o Excel em portugues abrir o arquivo direito
        with open(destino, "w", encoding="utf-8-sig", newline="") as arquivo:
            escritor = csv.writer(arquivo, delimiter=";")
            escritor.writerow(["Data", "Tipo", "Historico", "Valor", "Saldo"])
            for linha in self.visiveis:
                escritor.writerow(
                    [
                        linha.data,
                        linha.tipo,
                        linha.historico,
                        formatar_moeda(linha.valor),
                        formatar_moeda(linha.saldo, ""),
                    ]
                )

    def _gerar_xlsx(self, destino: str) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        planilha = Workbook()
        aba = planilha.active
        aba.title = "Extrato"

        cabecalhos = ["Data", "Tipo", "Historico", "Valor", "Saldo"]
        aba.append(cabecalhos)
        fundo = PatternFill("solid", fgColor="E8EEF5")
        for coluna in range(1, len(cabecalhos) + 1):
            celula = aba.cell(row=1, column=coluna)
            celula.font = Font(bold=True)
            celula.fill = fundo
            celula.alignment = Alignment(horizontal="center")

        for linha in self.visiveis:
            aba.append([linha.data, linha.tipo, linha.historico, linha.valor, linha.saldo])

        formato = 'R$ #,##0.00;[RED]-R$ #,##0.00'
        for fileira in aba.iter_rows(min_row=2, min_col=4, max_col=5):
            for celula in fileira:
                celula.number_format = formato

        for coluna, largura in zip(range(1, 6), (12, 16, 70, 16, 16)):
            aba.column_dimensions[get_column_letter(coluna)].width = largura

        aba.freeze_panes = "A2"
        aba.auto_filter.ref = f"A1:E{aba.max_row}"
        planilha.save(destino)

    # ------------------------------------------------------------------
    # Utilitarios
    # ------------------------------------------------------------------
    def _erro(self, titulo: str, mensagem: str) -> None:
        self.status.configure(text=titulo)
        messagebox.showerror(titulo, mensagem, parent=self.root)

    def _sucesso(self, destino: str, mensagem: str) -> None:
        self.status.configure(text=f"{mensagem} {destino}")
        if messagebox.askyesno("Concluido", f"{mensagem}\n\n{destino}\n\nDeseja abrir o arquivo agora?", parent=self.root):
            try:
                os.startfile(destino)  # type: ignore[attr-defined]
            except Exception:
                pass

    def _carregar_config(self) -> dict:
        try:
            with open(caminho_config(), encoding="utf-8") as arquivo:
                dados = json.load(arquivo)
                return dados if isinstance(dados, dict) else {}
        except Exception:
            return {}

    def _salvar_config(self) -> None:
        try:
            self.config["geometria"] = self.root.geometry()
            caminho = caminho_config()
            os.makedirs(os.path.dirname(caminho), exist_ok=True)
            with open(caminho, "w", encoding="utf-8") as arquivo:
                json.dump(self.config, arquivo, ensure_ascii=False, indent=2)
        except Exception:
            pass  # preferencias sao um conforto, nunca um impeditivo

    def _ao_fechar(self) -> None:
        self._salvar_config()
        self.root.destroy()


def main() -> None:
    root = tk.Tk()
    ConciliaApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
